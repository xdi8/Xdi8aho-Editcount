import openpyxl as xl
from requests import get, packages
from requests.exceptions import RequestException
from retrying import retry
from requests.packages.urllib3.exceptions import InsecureRequestWarning

import json
import os
import threading
from time import localtime


packages.urllib3.disable_warnings(InsecureRequestWarning)
__author__ = "QWERTY_52_38"
__version__ = "0.5"
rev_api = "https://wiki.xdi8.top/w/api.php?action=query&format=json&prop=revisions&revids="
# revisions api
folder = r"D:\python\xdi8\Xdi8aho-Editcount"  # this should be changed to your own directory

namespace_score = {0: 3, 1: 0.125, 4: 1, 5: 0.125,
                   10: 4, 11: 0.125, 12: 3, 13: 0.125, 14: 1, 15: 0.125,
                   3824: 0.5, 3825: 0.125, 3826: 0.5, 3827: 0.125}

namespace_loca = {0: 0, 1: 1, 3: 1, 4: 2, 5: 1, 10: 3, 
                   11: 1, 12: 4, 13: 1, 14: 5, 15: 1,
                   3824: 6, 3825: 1, 3826: 7, 3827: 1}


@retry(stop_max_attempt_number=10)
def get_page(url: str):
    return get(url, timeout=15, verify=False).text


def get_revs(start, end):
    for i in range(start, end + 1):
        pth = os.path.join(os.path.join(folder, "rev"), f"rev_{i}.txt")
        if os.path.isfile(pth):
            continue
        rev = get_page(rev_api + str(i))
        with open(pth, "w") as fh:
            fh.write(rev)
    print(f"{start} to {end} finished!")


def get_edit_score_dic(start: int, end: int) -> dict:
    user_dic = {}
    for i in range(start, end + 1):
        with open(os.path.join(os.path.join(folder, "rev"), f"rev_{i}.txt"), "r") as fh:
            try:
                js = json.loads(fh.read())
                page_id = list(js["query"]["pages"].keys())[0]
                namespace = js["query"]["pages"][page_id]["ns"]
                title = js["query"]["pages"][page_id]["title"]
                user = js["query"]["pages"][page_id]["revisions"][0]["user"]
            except:
                continue
            if user not in user_dic:
                user_dic[user] = [0,0,0,0,0,0,0,0,0,0]  # main talk wiki template help category fun word score total
            if namespace in namespace_score:
                user_dic[user][namespace_loca[namespace]] += 1
                user_dic[user][-2] += namespace_score[namespace]
            user_dic[user][-1] += 1
    return user_dic


def make_workbook(dic: dict, filename=f"xdi8aho-wiki-useredit-{localtime().tm_year}{localtime().tm_mon}{localtime().tm_mday}-QWERTY770.xlsx"):
    wb = xl.Workbook()
    ws = wb.create_sheet('main',0)
    
    ws.cell(row=1, column=1).value = "用户名"
    ws.cell(row=1, column=2).value = "编辑总计"
    ws.cell(row=1, column=3).value = "（主）"
    ws.cell(row=1, column=4).value = "讨论"
    ws.cell(row=1, column=5).value = "希顶维基"
    ws.cell(row=1, column=6).value = "模板"
    ws.cell(row=1, column=7).value = "帮助"
    ws.cell(row=1, column=8).value = "分类"
    ws.cell(row=1, column=9).value = "生草"
    ws.cell(row=1, column=10).value = "词汇"
    ws.cell(row=1, column=11).value = "编辑积分"

    for m, i in enumerate(dic.keys()):
        ws.cell(row=m+2, column=1).value = i
        ws.cell(row=m+2, column=2).value = dic[i][-1]
        for n, j in enumerate(dic[i]):
            if n != len(dic[i]) - 1:
                ws.cell(row=m+2, column=n+3).value = j
    
    wb.save(os.path.join(folder, filename))
    wb.close()


if __name__ == "__main__":
    '''
    thread_list = []
    for i in range(8):
        t = threading.Thread(target=get_revs, args=(2913*i, 2913+2913*i))
        t.start()
        thread_list.append(t)
    for j in thread_list:
        j.join()
    '''
    get_revs(23957, 24362)
    make_workbook(get_edit_score_dic(1, 24362))
    print("Finished!")
